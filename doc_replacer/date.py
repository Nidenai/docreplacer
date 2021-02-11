# Тут написаны функции и переменные, касающиеся дат, которые в данный момент берутся из книги Excel для последующего добавления в документ
import openpyxl

# создал вспомогательную функцию для добавления нолика перед числом, по типу 01.02

def is_zero(number):
    if number < 10:
        return '0' + str(number)
    return number

wb = openpyxl.load_workbook('ref/reference_book.xlsx', data_only=True)
ws = wb.active
day_begin = is_zero(ws['G2'].value)
month_begin = is_zero(ws['H2'].value)
day_end = is_zero(ws['G3'].value)
month_end = is_zero(ws['H3'].value)

def is_month(month):
    if month == '01':
        date_month = 'января'
    elif month == '02':
        date_month = 'февраля'
    elif month == '03':
        date_month = 'марта'
    elif month == '04':
        date_month = 'апреля'
    elif month == '05':
        date_month = 'мая'
    elif month == '06':
        date_month = 'июня'
    elif month == '07':
        date_month = 'июля'
    elif month == '08':
        date_month = 'августа'
    elif month == '09':
        date_month = 'сентября'
    elif month == '10':
        date_month = 'октября'
    elif month == '11':
        date_month = 'ноября'
    elif month == '12':
        date_month = 'декабря'
    return date_month

string_month = is_month(month_begin)

print(day_begin, day_end, month_begin, month_end, string_month)